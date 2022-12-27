import openpyxl
loadxl = openpyxl.load_workbook("inventory.xlsx")
product_list = loadxl['Sheet1']
####### Ex:1 List each company with respective product count
## create a empty dictionary to store supplier name

def get_product_count_each_compant():
    supplier_name_list = {}
    #print(product_list.max_row)

    for product_row in range(2,product_list.max_row + 1):
        supplier_name = product_list.cell(product_row,4).value
        if supplier_name in supplier_name_list:
            supplier_name_list[supplier_name] = supplier_name_list.get(supplier_name) + 1
        else:
            print("adding new supplier")
            supplier_name_list[supplier_name] = 1        

    print(supplier_name_list)   

#get_product_count_each_compant()
###### List products with inventory less than 10

def list_product_lessThan_ten():

    product_name_list_ten = {}

    for product_row in range(2,product_list.max_row + 1):
   
        inventory_number = product_list.cell(product_row,2).value
        #print(product_row, inventory_number) 
        if inventory_number < 10 : 
            product_no = int(product_list.cell(product_row,1).value)
            product_name_list_ten[product_no] = int(inventory_number) 
    return product_name_list_ten
#print(list_product_lessThan_ten())

###### Calculate each comapany with respective local inventory value
def calculate_product_name_with_value():
    product_name_with_value = {}
    #supplier_name_list = {}
    for product_row in range(2,product_list.max_row + 1):
        supplier_name = product_list.cell(product_row,4).value
        inventory_count = product_list.cell(product_row,2).value
        inventory_price = product_list.cell(product_row,3).value
        
        if supplier_name in product_name_with_value:
            product_name_with_value[supplier_name] = product_name_with_value[supplier_name] + int(inventory_count * inventory_price)
        else:
            
            product_name_with_value[supplier_name] = int(inventory_count * inventory_price)       
    return product_name_with_value

###### Write to spreasheet:  Calculate and write inventory value for each product into spreadsheet
def calculate_product_value_each_company():
    product_name_with_value = {}
    ## set a new column name for calculate_product_value_each_company
    calculate_product_value_each_company = product_list.cell(1,5)
    calculate_product_value_each_company.value = 'Total'
    
    for product_row in range(2,product_list.max_row + 1):
        supplier_name = product_list.cell(product_row,4).value
        inventory_count = product_list.cell(product_row,2).value
        inventory_price = product_list.cell(product_row,3).value
        calculate_product_value_each_company = product_list.cell(product_row,5)

        calculate_product_value_each_company.value = inventory_count * inventory_price
    
    loadxl.save('inventory_with_total_value.xlsx')

calculate_product_value_each_company()

        
 
